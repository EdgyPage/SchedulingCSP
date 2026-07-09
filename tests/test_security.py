import io

import pytest
from openpyxl import load_workbook
from fastapi import HTTPException
from fastapi.testclient import TestClient

import config
import employee as e
import io_adapters as ia
from app import app, _validate_inputs

VALID_PAYLOAD = {
    "tasks": ["A", "B"],
    "employees": [
        {"id": 1, "name": "Alice", "approved": ["A"]},
        {"id": 2, "name": "Bob", "approved": ["B"]},
    ],
    "minimums": [1, 1],
    "max_schedules": 2,
    "seed": 0,
}


# --- formula / CSV injection -------------------------------------------------

def test_json_is_not_sanitized_but_exports_are():
    emp = e.Employee("=id", "=name", ["A"], [True])
    internal = [{"A": [emp], "Unassigned": []}]

    tables = ia.schedules_to_json(internal)
    assert tables[0][0]["Name"] == "=name"   # JSON stays raw (frontend escapes)
    assert tables[0][0]["ID"] == "=id"

    csv = ia.results_to_csv_bytes(tables)
    assert b"'=name" in csv and b"'=id" in csv

    xlsx = ia.results_to_xlsx_bytes(tables)
    ws = load_workbook(io.BytesIO(xlsx))["Schedule_1"]
    header = [cell.value for cell in ws[1]]
    first = dict(zip(header, [cell.value for cell in ws[2]]))
    assert first["Name"] == "'=name"


@pytest.mark.parametrize("dangerous", ["=1+1", "+1", "-1", "@SUM(A1)", "\tx", "\rx"])
def test_sanitize_cell_neutralizes_triggers(dangerous):
    assert ia._sanitize_cell(dangerous) == "'" + dangerous


def test_sanitize_cell_leaves_safe_values():
    assert ia._sanitize_cell("Alice") == "Alice"
    assert ia._sanitize_cell(42) == 42  # non-strings untouched


def test_export_endpoint_sanitizes(monkeypatch):
    monkeypatch.setattr(config, "API_KEY", None)
    client = TestClient(app)
    schedules = [[{"Name": "=cmd|'/C calc'!A1", "ID": 1, "Function": "A"}]]
    r = client.post("/api/export?format=csv", json={"schedules": schedules})
    assert r.status_code == 200
    assert b"'=cmd" in r.content


# --- resource limits / validation -------------------------------------------

def test_config_clamps_budget_and_schedules():
    assert config.effective_time_budget(None) == config.MAX_TIME_BUDGET_S
    assert config.effective_time_budget(10 ** 9) == config.MAX_TIME_BUDGET_S
    assert config.effective_time_budget(-5) == config.MAX_TIME_BUDGET_S
    small = min(1.0, config.MAX_TIME_BUDGET_S)
    assert config.effective_time_budget(small) == small
    assert config.effective_max_schedules(10 ** 9) == config.MAX_SCHEDULES
    assert config.effective_max_schedules(0) == 1


def test_validate_inputs_rejects_bad_minimums():
    with pytest.raises(HTTPException):
        _validate_inputs([], ["A"], [float("nan")], 1)      # NaN
    with pytest.raises(HTTPException):
        _validate_inputs([], ["A"], [-1], 1)                # negative
    with pytest.raises(HTTPException):
        _validate_inputs([], ["A", "B"], [1], 1)            # wrong length


def test_oversized_roster_and_tasks_rejected(monkeypatch):
    client = TestClient(app)
    monkeypatch.setattr(config, "MAX_EMPLOYEES", 1)
    payload = {
        "tasks": ["A"],
        "employees": [
            {"id": 1, "name": "a", "approved": ["A"]},
            {"id": 2, "name": "b", "approved": ["A"]},
        ],
        "minimums": [1],
    }
    assert client.post("/api/solve", json=payload).status_code == 400

    monkeypatch.setattr(config, "MAX_TASKS", 1)
    assert client.post(
        "/api/solve",
        json={"tasks": ["A", "B"], "employees": [], "minimums": [1, 1]},
    ).status_code == 400


def test_oversized_upload_rejected(monkeypatch):
    monkeypatch.setattr(config, "MAX_UPLOAD_BYTES", 5)
    client = TestClient(app)
    csv = b"ID,Name,A\n1,alice,1\n"
    r = client.post(
        "/api/solve/file",
        files={"file": ("roster.csv", csv, "text/csv")},
        data={"minimums": "[1]"},
    )
    assert r.status_code == 413


# --- optional API-key gate ---------------------------------------------------

def test_api_key_gate(monkeypatch):
    client = TestClient(app)

    monkeypatch.setattr(config, "API_KEY", None)
    assert client.post("/api/solve", json=VALID_PAYLOAD).status_code == 200

    monkeypatch.setattr(config, "API_KEY", "s3cret")
    assert client.post("/api/solve", json=VALID_PAYLOAD).status_code == 401
    assert client.post(
        "/api/solve", json=VALID_PAYLOAD, headers={"X-API-Key": "s3cret"}
    ).status_code == 200


def test_happy_path_still_works(monkeypatch):
    monkeypatch.setattr(config, "API_KEY", None)
    client = TestClient(app)
    r = client.post("/api/solve", json=VALID_PAYLOAD)
    assert r.status_code == 200
    body = r.json()
    assert body["count"] >= 1
    assert body["tasks"] == ["A", "B"]
