"""Structural checks for the roster-template scaffold builder.

Excel formulas and Power Query can't run in this environment, so this only asserts the
workbook STRUCTURE (sheets, tables, headers, and the Decoder XLOOKUP formula strings). The
encode/decode behavior itself must be verified in Excel -- see docs/roster-template-spec.md.
"""

import importlib.util
import os

import pytest
from openpyxl import load_workbook

_HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_BUILDER = os.path.join(_HERE, "frontend", "build_roster_template.py")


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    spec = importlib.util.spec_from_file_location("build_roster_template", _BUILDER)
    builder = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(builder)
    out = tmp_path_factory.mktemp("tpl") / "roster-template.xlsx"
    builder.build(str(out))
    return load_workbook(out)


def _headers(ws, table_name):
    return [cell.value for cell in ws[ws.tables[table_name].ref][0]]


def test_sheets_and_hidden_maps(wb):
    assert wb.sheetnames == ["Encoder", "Encoded", "Maps", "Decoder"]
    assert wb["Maps"].sheet_state == "hidden"


def test_tables_and_headers(wb):
    assert _headers(wb["Encoder"], "Roster")[0] == "Unique Identifier"
    assert _headers(wb["Encoded"], "Encoded") == ["ID Alias", "Func 1", "Func 2", "Func 3"]
    assert _headers(wb["Maps"], "Master") == \
        ["ID Alias", "Unique Identifier", "Func 1", "Func 2", "Func 3"]
    assert _headers(wb["Maps"], "FuncMap") == ["Func Code", "Function Name"]
    assert _headers(wb["Decoder"], "SolverOutput") == \
        ["ID Alias", "Function", "Unique Identifier", "Function Name"]


def test_decoder_xlookup_formulas(wb):
    dec = wb["Decoder"]
    id_formula = dec["C3"].value      # first data row of the two calculated columns
    func_formula = dec["D3"].value
    assert id_formula.startswith("=_xlfn.XLOOKUP(") and "Master[ID Alias]" in id_formula
    assert func_formula.startswith("=_xlfn.XLOOKUP(") and "FuncMap[Func Code]" in func_formula
    # Unassigned (and any unmapped Function) passes through via XLOOKUP's if_not_found arg.
    assert func_formula.endswith("[@Function])")


def test_committed_template_matches_builder(wb):
    """The committed frontend/roster-template.xlsx should have the same structure as the builder."""
    committed = os.path.join(_HERE, "frontend", "roster-template.xlsx")
    if not os.path.exists(committed):
        pytest.skip("template not generated/committed yet")
    cwb = load_workbook(committed)
    assert cwb.sheetnames == wb.sheetnames
    assert _headers(cwb["Encoded"], "Encoded") == _headers(wb["Encoded"], "Encoded")
