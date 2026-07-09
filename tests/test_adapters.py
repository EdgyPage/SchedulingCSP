import csv
import io

import pytest
from openpyxl import load_workbook

import io_adapters as ia

ROWS = [
    {"ID Alias": 1, "Func 1": True, "Func 2": False},
    {"ID Alias": 2, "Func 1": False, "Func 2": True},
]
COLUMNS = ["ID Alias", "Func 1", "Func 2"]


def _rows_to_csv_bytes(rows, columns):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns)
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode()


def _read_xlsx_sheet(xlsx, sheet):
    wb = load_workbook(io.BytesIO(xlsx))
    rows = list(wb[sheet].iter_rows(values_only=True))
    header = list(rows[0])
    data = [dict(zip(header, row)) for row in rows[1:]]
    return header, data


def test_parsers_agree_across_formats():
    emps_rec, tasks_rec = ia.parse_roster_records(ROWS, COLUMNS)

    emps_csv, tasks_csv = ia.parse_roster_csv(_rows_to_csv_bytes(ROWS, COLUMNS))

    form = {
        "tasks": ["Func 1", "Func 2"],
        "employees": [
            {"id": 1, "approved": ["Func 1"]},
            {"id": 2, "approved": ["Func 2"]},
        ],
    }
    emps_form, tasks_form = ia.parse_roster_form(form)

    assert tasks_rec == tasks_csv == tasks_form == ["Func 1", "Func 2"]
    for group in (emps_rec, emps_csv, emps_form):
        assert [emp.taskStatuses for emp in group] == [[True, False], [False, True]]


def test_name_or_pii_column_is_rejected():
    # A Name (or any non-'Func' column) must be refused so PII can't be uploaded.
    with pytest.raises(ValueError, match="Unexpected column"):
        ia.parse_roster_records(
            [{"ID Alias": 1, "Func 1": True, "Name": "alice"}],
            ["ID Alias", "Func 1", "Name"],
        )


def test_result_serializers_produce_valid_files():
    results = [
        [
            {"ID Alias": 1, "Function": "Func 1"},
            {"ID Alias": 2, "Function": "Unassigned"},
        ],
    ]

    xlsx = ia.results_to_xlsx_bytes(results)
    assert xlsx[:2] == b"PK"  # xlsx is a zip archive
    header, data = _read_xlsx_sheet(xlsx, "Schedule_1")
    assert header == ["ID Alias", "Function"]
    assert len(data) == 2

    csv_bytes = ia.results_to_csv_bytes(results)
    assert b"Schedule,ID Alias,Function" in csv_bytes


def test_closest_export_includes_summary_when_meta_given():
    """A near-miss export carries a Summary sheet / block; a plain export does not."""
    results = [[{"ID Alias": 1, "Function": "Func 1"}]]
    meta = {
        "tasks": ["Func 1", "Func 2"], "target": [2, 1], "metric": "cosine", "mode": "thorough",
        "schedules": [
            {"distance": 0.05, "coverage": [1, 1], "shortfall": [1, 0],
             "covered": 2, "target_total": 3},
        ],
    }

    xlsx = ia.results_to_xlsx_bytes(results, meta=meta)
    assert "Summary" in load_workbook(io.BytesIO(xlsx)).sheetnames
    assert "Summary" not in load_workbook(io.BytesIO(ia.results_to_xlsx_bytes(results))).sheetnames

    csv_bytes = ia.results_to_csv_bytes(results, meta=meta)
    assert b"Closest schedules" in csv_bytes
    assert b"Closest schedules" not in ia.results_to_csv_bytes(results)
