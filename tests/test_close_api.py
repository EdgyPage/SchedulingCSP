import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import config
from app import app

INFEASIBLE = {
    "tasks": ["Func 1", "Func 2", "Func 3"],
    "employees": [
        {"id": 1, "approved": ["Func 1"]},
        {"id": 2, "approved": ["Func 2"]},
        {"id": 3, "approved": ["Func 3"]},
    ],
    "minimums": [1, 1, 3],       # Func 3 needs 3 but only employee 3 can do it
    "seed": 0,
}


def _client(monkeypatch):
    monkeypatch.setattr(config, "API_KEY", None)
    return TestClient(app)


def test_solve_endpoint_returns_infeasibility_and_closest(monkeypatch):
    body = _client(monkeypatch).post("/api/solve", json=INFEASIBLE).json()
    assert body["count"] == 0
    assert body["tasks"] == ["Func 1", "Func 2", "Func 3"]
    assert any(r["task"] == "Func 3" for r in body["infeasibility"]["reasons"])
    closest = body["closest"]
    assert closest["metric"] == "cosine" and closest["mode"] == "thorough"
    assert closest["schedules"][0]["coverage"] == [1, 1, 1]


def test_solve_endpoint_rejects_bad_metric_mode_and_minimums(monkeypatch):
    client = _client(monkeypatch)
    assert client.post("/api/solve", json={**INFEASIBLE, "metric": "nope"}).status_code == 400
    assert client.post("/api/solve", json={**INFEASIBLE, "mode": "nope"}).status_code == 400
    assert client.post("/api/solve", json={**INFEASIBLE, "minimums": [1, 1, 1.5]}).status_code == 400


def test_export_with_meta_adds_summary(monkeypatch):
    client = _client(monkeypatch)
    payload = {
        "schedules": [[{"ID Alias": 1, "Function": "Func 1"}]],
        "meta": {
            "tasks": ["Func 1", "Func 2"], "target": [1, 2], "metric": "cosine", "mode": "thorough",
            "schedules": [{"distance": 0.1, "coverage": [1, 1], "shortfall": [0, 1],
                           "covered": 2, "target_total": 3}],
        },
    }
    xlsx = client.post("/api/export?format=xlsx", json=payload)
    assert xlsx.status_code == 200
    assert "Summary" in load_workbook(io.BytesIO(xlsx.content)).sheetnames
    assert b"Closest schedules" in client.post("/api/export?format=csv", json=payload).content
