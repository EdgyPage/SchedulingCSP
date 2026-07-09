"""Input parsers and output serializers that bridge the pure solver to the web layer.

Inputs (Excel upload, CSV upload, or a web form) all converge to the same internal
shape: ``(list[Employee], list[task_name])``. Outputs are produced in-memory (no disk
writes) as JSON tables, an Excel workbook, or a CSV -- suitable for returning directly
from an HTTP handler.

Expected roster layout (xlsx/csv): an ``ID`` column, a ``Name`` column, and one column
per task whose cell is truthy when the employee is approved for that task.

Spreadsheet/CSV I/O uses ``openpyxl`` and the stdlib ``csv`` module -- no pandas.
"""

import csv
import io
from datetime import datetime

from openpyxl import Workbook, load_workbook

from employee import Employee

ID_COL = "ID"
NAME_COL = "Name"
_OUTPUT_COLUMNS = ["Name", "ID", "Function"]


# --- helpers ----------------------------------------------------------------

def _to_bool(value) -> bool:
    """Normalize a spreadsheet/form cell to a real bool (handles None, 1/0, 'x', etc.)."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value != value:  # NaN
            return False
        return value != 0
    return str(value).strip().lower() in {
        "true", "t", "yes", "y", "1", "x", "approved", "checked",
    }


def _coerce_id(value):
    """Best-effort numeric ID. openpyxl yields real types; stdlib csv yields strings,
    so an ID column of ``1, 2, 3`` comes back as ``"1", "2", "3"`` -- cast those back to
    int so ``Employee.id`` stays numeric as it did under pandas' type inference."""
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = str(value).strip()
    try:
        return int(text)
    except ValueError:
        return value


def _approved_to_statuses(approved, task_names) -> list[bool]:
    """Accept approvals as a name-subset list, a flag list, or a {task: flag} mapping."""
    if isinstance(approved, dict):
        return [_to_bool(approved.get(t, False)) for t in task_names]
    approved = list(approved)
    all_strings = all(isinstance(a, str) for a in approved)
    if not all_strings and len(approved) == len(task_names):
        return [_to_bool(a) for a in approved]
    approved_set = {str(a) for a in approved}
    return [t in approved_set for t in task_names]


# --- input parsers ----------------------------------------------------------

def parse_roster_records(records: list[dict], columns: list[str]):
    """Shared parser: turn ordered ``columns`` + row ``records`` (header->cell dicts)
    into ``(list[Employee], list[task_name])``. Used by both the xlsx and csv readers."""
    missing = [col for col in (ID_COL, NAME_COL) if col not in columns]
    if missing:
        raise ValueError(f"Roster is missing required column(s): {', '.join(missing)}")
    task_names = [col for col in columns if col not in (ID_COL, NAME_COL)]
    if not task_names:
        raise ValueError("Roster has no task columns.")
    employees = []
    for row in records:
        raw_id = row.get(ID_COL)
        # Skip fully-blank rows (openpyxl read_only can emit trailing empties).
        if raw_id is None or (isinstance(raw_id, str) and not raw_id.strip()):
            continue
        statuses = [_to_bool(row.get(task)) for task in task_names]
        employees.append(
            Employee(_coerce_id(raw_id), str(row.get(NAME_COL, "")), list(task_names), statuses)
        )
    return employees, task_names


def _worksheet_to_records(ws):
    """First row is the header; remaining rows become header->cell dicts."""
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []
    columns = [str(h).strip() for h in header_row if h is not None]
    records = []
    for row in rows:
        if row is None or all(v is None for v in row):
            continue
        records.append({col: (row[i] if i < len(row) else None) for i, col in enumerate(columns)})
    return records, columns


def parse_roster_xlsx(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        records, columns = _worksheet_to_records(wb.active)
    finally:
        wb.close()
    return parse_roster_records(records, columns)


def parse_roster_csv(file_bytes: bytes):
    # utf-8-sig tolerates a BOM some spreadsheet exporters prepend.
    reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig")))
    columns = list(reader.fieldnames or [])
    records = [dict(row) for row in reader]
    return parse_roster_records(records, columns)


def parse_roster_form(payload: dict):
    task_names = list(payload["tasks"])
    if not task_names:
        raise ValueError("At least one task is required.")
    employees = []
    for entry in payload["employees"]:
        statuses = _approved_to_statuses(entry.get("approved", []), task_names)
        employees.append(Employee(entry["id"], str(entry["name"]), list(task_names), statuses))
    return employees, task_names


# --- output serializers -----------------------------------------------------

# Cells beginning with one of these are interpreted as a formula by Excel/Sheets.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value):
    """Neutralize spreadsheet formula injection (OWASP CSV injection).

    Employee names, IDs, and task names are user-supplied and end up in downloadable
    files. If a string cell would be read as a formula, prefix it with a single quote
    so the spreadsheet treats it as text. Non-string cells pass through unchanged.
    """
    if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _sanitize_table(table: list[dict]) -> list[dict]:
    return [{key: _sanitize_cell(val) for key, val in row.items()} for row in table]


def _schedule_to_rows(assignment: dict) -> list[dict]:
    rows = []
    for task, employees in assignment.items():
        for employee in employees:
            rows.append({"Name": employee.name, "ID": employee.id, "Function": task})
    return rows


def schedules_to_json(schedules: list[dict]) -> list[list[dict]]:
    """Convert internal ``{task: [Employee]}`` schedules to JSON-serializable tables.

    Not sanitized on purpose: JSON is safe structured data, and escaping for display
    is the frontend's responsibility. Only the file exports below are sanitized.
    """
    return [_schedule_to_rows(assignment) for assignment in schedules]


def _summary_rows(meta: dict) -> list[list]:
    """Build the human-readable near-miss summary block shared by the xlsx/csv exports.

    ``meta`` (optional) describes 'closest' results: ``tasks``, ``target``, ``metric``,
    ``mode``, and per-schedule ``distance``/``coverage``/``shortfall``/``covered``/
    ``target_total``. Returns a list of rows (each a list of cells)."""
    tasks = list(meta.get("tasks", []))
    target = list(meta.get("target", []))
    rows = [[f"Closest schedules (metric: {meta.get('metric', '?')}, mode: {meta.get('mode', '?')})"]]
    rows.append(["Schedule", "Distance", "Covered", "Target total", *[str(t) for t in tasks]])
    for i, sched in enumerate(meta.get("schedules", []), start=1):
        coverage = sched.get("coverage", [])
        shortfall = sched.get("shortfall", [])
        per_task = []
        for k in range(len(tasks)):
            cov = coverage[k] if k < len(coverage) else ""
            tgt = target[k] if k < len(target) else ""
            short = shortfall[k] if k < len(shortfall) else 0
            per_task.append(f"{cov}/{tgt}" + (f" (short {short})" if short else ""))
        rows.append([f"Closest {i}", sched.get("distance"), sched.get("covered"),
                     sched.get("target_total"), *per_task])
    return rows


def results_to_xlsx_bytes(results: list[list[dict]], meta: dict | None = None) -> bytes:
    """One workbook, one sheet per schedule. ``results`` is the schedules_to_json shape.

    When ``meta`` is given (a 'closest'/near-miss result), a leading ``Summary`` sheet
    describes how close each schedule is; without it, output matches the exact-schedule
    exports.
    """
    wb = Workbook()
    default = wb.active  # the auto-created empty sheet; removed once real sheets exist
    if meta:
        summary = wb.create_sheet("Summary")
        for row in _summary_rows(meta):
            summary.append([_sanitize_cell(cell) for cell in row])
    if not results:
        wb.create_sheet("Schedule_1").append(_OUTPUT_COLUMNS)
    for i, table in enumerate(results, start=1):
        ws = wb.create_sheet(f"Schedule_{i}")
        ws.append(_OUTPUT_COLUMNS)
        for row in _sanitize_table(table):
            ws.append([row.get(col) for col in _OUTPUT_COLUMNS])
    wb.remove(default)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def results_to_csv_bytes(results: list[list[dict]], meta: dict | None = None) -> bytes:
    """Single CSV; a leading ``Schedule`` column distinguishes each schedule.

    When ``meta`` is given, a labeled summary block precedes the assignment rows.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    if meta:
        for row in _summary_rows(meta):
            writer.writerow([_sanitize_cell(cell) for cell in row])
        writer.writerow([])  # blank line separates the summary from the assignments
    writer.writerow(["Schedule"] + _OUTPUT_COLUMNS)
    for i, table in enumerate(results, start=1):
        for row in _sanitize_table(table):
            writer.writerow([i] + [row.get(col) for col in _OUTPUT_COLUMNS])
    return buffer.getvalue().encode("utf-8")


def write_schedules_to_excel_files(schedules: list[dict], filePath: str = "") -> list[str]:
    """Legacy/local convenience: write one .xlsx file per schedule to disk."""
    paths = []
    for i, assignment in enumerate(schedules, start=1):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = f"{filePath}_Schedule_{i}_{timestamp}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = f"Schedule_{i}"
        ws.append(_OUTPUT_COLUMNS)
        for row in _sanitize_table(_schedule_to_rows(assignment)):
            ws.append([row.get(col) for col in _OUTPUT_COLUMNS])
        wb.save(path)
        paths.append(path)
    return paths
