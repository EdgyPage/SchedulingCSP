"""Generate ``frontend/roster-template.xlsx`` -- the de-identified encode/decode scaffold.

This builds the STRUCTURE non-technical users download: the four sheets (Encoder, Encoded,
Maps, Decoder), the input/lookup Tables, seeded example data, and the Decoder's XLOOKUP
formulas. The one-click encode step is added in Excel as a single Power Query -- openpyxl
cannot author Power Query, so see ``docs/roster-template-spec.md`` for the M code and setup.
The Encoded/Maps sheets carry example values here that the query overwrites on ``Refresh All``.

Run:  python frontend/build_roster_template.py
"""

import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# Example roster (users replace these with their real identifiers + real function names).
FUNCS = ["Morning Shift", "Register", "Stocking"]
EXAMPLE = [
    ("E-1042", [True, False, True]),
    ("E-2087", [False, True, True]),
    ("E-3310", [True, True, False]),
]
ALIAS = [2, 3, 1]  # a fixed example permutation for the preview; the query randomizes for real

# XLOOKUP formulas: the '_xlfn.' prefix is required in the stored file; Excel shows "XLOOKUP".
# Both look up the single Master table + FuncMap, which the one Power Query keeps in sync.
ID_FORMULA = ('=_xlfn.XLOOKUP([@[ID Alias]],Master[ID Alias],'
              'Master[Unique Identifier],"(unknown alias)")')
FUNC_FORMULA = ('=_xlfn.XLOOKUP([@Function],FuncMap[Func Code],'
                'FuncMap[Function Name],[@Function])')


def _add_table(ws, name, first_row, ncols, ndata):
    ref = f"A{first_row}:{get_column_letter(ncols)}{first_row + ndata}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)


def build(path):
    wb = Workbook()

    # --- Encoder: users paste REAL data here -------------------------------
    enc = wb.active
    enc.title = "Encoder"
    enc["A1"] = ("STEP 1 - Replace the example rows with your real employees and rename the "
                 "function columns to your real functions. Mark each cell TRUE or FALSE, then "
                 "Data > Refresh All (after adding the Encode query -- see the Instructions).")
    enc.append(["Unique Identifier", *FUNCS])          # header -> row 2
    for ident, approvals in EXAMPLE:
        enc.append([ident, *approvals])
    _add_table(enc, "Roster", first_row=2, ncols=1 + len(FUNCS), ndata=len(EXAMPLE))
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    enc.add_data_validation(dv)
    dv.add(f"B3:{get_column_letter(1 + len(FUNCS))}2000")

    # --- Encoded: upload-ready, de-identified (the query fills this) --------
    encoded = wb.create_sheet("Encoded")
    encoded["A1"] = ("Auto-filled by Refresh All. Copy this whole table into a blank file, Save "
                     "As .csv, and upload it on the Solver page. (Example values shown for now.)")
    encoded.append(["ID Alias", *[f"Func {i}" for i in range(1, len(FUNCS) + 1)]])
    for alias, (_, approvals) in sorted(zip(ALIAS, EXAMPLE)):
        encoded.append([alias, *approvals])
    _add_table(encoded, "Encoded", first_row=2, ncols=1 + len(FUNCS), ndata=len(EXAMPLE))

    # --- Maps (hidden): the single Master table + FuncMap the Decoder uses --
    maps = wb.create_sheet("Maps")
    maps.sheet_state = "hidden"
    maps["A1"] = "Auto-filled by Refresh All. Do not edit."
    # Master is the one source of truth the Power Query writes: alias + real identifier +
    # generic funcs. Encoded (upload) and the Decoder both derive from it, so they can't drift.
    maps.append(["ID Alias", "Unique Identifier", *[f"Func {i}" for i in range(1, len(FUNCS) + 1)]])
    for alias, (ident, approvals) in sorted(zip(ALIAS, EXAMPLE)):
        maps.append([alias, ident, *approvals])
    _add_table(maps, "Master", first_row=2, ncols=2 + len(FUNCS), ndata=len(EXAMPLE))
    # FuncMap in a disjoint block (leave a gap after Master): Func code -> real function name.
    fm = 2 + len(FUNCS) + 2
    c1, c2 = get_column_letter(fm), get_column_letter(fm + 1)
    maps[f"{c1}2"], maps[f"{c2}2"] = "Func Code", "Function Name"
    for i, fname in enumerate(FUNCS, start=1):
        maps[f"{c1}{2 + i}"], maps[f"{c2}{2 + i}"] = f"Func {i}", fname
    fmap = Table(displayName="FuncMap", ref=f"{c1}2:{c2}{2 + len(FUNCS)}")
    fmap.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    maps.add_table(fmap)

    # --- Decoder: paste solver output, XLOOKUPs fill real identity ---------
    dec = wb.create_sheet("Decoder")
    dec["A1"] = ("STEP 2 - Paste the solver's two columns (ID Alias, Function) into the first two "
                 "columns below. The real identifier and function name fill in automatically.")
    dec.append(["ID Alias", "Function", "Unique Identifier", "Function Name"])
    for alias, func in [(1, "Func 1"), (2, "Unassigned")]:
        dec.append([alias, func, ID_FORMULA, FUNC_FORMULA])
    _add_table(dec, "SolverOutput", first_row=2, ncols=4, ndata=2)

    wb.save(path)


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "roster-template.xlsx")
    build(out)
    print(f"wrote {out}")
